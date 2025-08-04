import pandas as pd
import aiofiles
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import io

class ExcelService:
    """Excel处理服务"""
    
    def __init__(self):
        self.output_dir = Path("outputs")
        self.output_dir.mkdir(exist_ok=True)
        self.template_dir = Path("templates")
        self.template_dir.mkdir(exist_ok=True)
    
    async def read_excel_file(self, file_path: Path) -> Dict[str, Any]:
        """
        读取Excel文件
        
        Args:
            file_path: Excel文件路径
        
        Returns:
            包含数据和元信息的字典
        """
        try:
            if not file_path.exists():
                raise FileNotFoundError(f"文件不存在: {file_path}")
            
            # 读取Excel文件
            df = pd.read_excel(file_path, sheet_name=None)
            
            result = {
                "filename": file_path.name,
                "sheets": {},
                "metadata": {
                    "total_sheets": len(df),
                    "sheet_names": list(df.keys()),
                    "file_size": file_path.stat().st_size,
                    "last_modified": datetime.fromtimestamp(file_path.stat().st_mtime).isoformat()
                }
            }
            
            # 处理每个工作表
            for sheet_name, sheet_df in df.items():
                # 将NaN值替换为空字符串
                sheet_df = sheet_df.fillna('')
                
                result["sheets"][sheet_name] = {
                    "data": sheet_df.to_dict('records'),
                    "columns": list(sheet_df.columns),
                    "shape": sheet_df.shape,
                    "dtypes": {col: str(dtype) for col, dtype in sheet_df.dtypes.items()}
                }
            
            return result
            
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {str(e)}")
    
    async def save_excel_data(
        self, 
        data: List[Dict[str, Any]], 
        filename: str,
        sheet_name: str = "Sheet1",
        output_dir: Optional[str] = None
    ) -> Path:
        """
        保存数据为Excel文件
        
        Args:
            data: 要保存的数据
            filename: 文件名
            sheet_name: 工作表名称
            output_dir: 输出目录
        
        Returns:
            保存的文件路径
        """
        try:
            # 确定输出目录
            save_dir = Path(output_dir) if output_dir else self.output_dir
            save_dir.mkdir(exist_ok=True)
            
            # 确保文件名有正确的扩展名
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            file_path = save_dir / filename
            
            # 将数据转换为DataFrame
            df = pd.DataFrame(data)
            
            # 保存为Excel文件
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 获取工作表对象进行格式化
                worksheet = writer.sheets[sheet_name]
                
                # 自动调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # 设置标题行样式
                if len(df) > 0:
                    for cell in worksheet[1]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            return file_path
            
        except Exception as e:
            raise Exception(f"保存Excel文件失败: {str(e)}")
    
    async def save_multiple_sheets(
        self, 
        sheets_data: Dict[str, List[Dict[str, Any]]], 
        filename: str,
        output_dir: Optional[str] = None
    ) -> Path:
        """
        保存多个工作表到一个Excel文件
        
        Args:
            sheets_data: 工作表数据字典 {sheet_name: data}
            filename: 文件名
            output_dir: 输出目录
        
        Returns:
            保存的文件路径
        """
        try:
            # 确定输出目录
            save_dir = Path(output_dir) if output_dir else self.output_dir
            save_dir.mkdir(exist_ok=True)
            
            # 确保文件名有正确的扩展名
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            file_path = save_dir / filename
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, data in sheets_data.items():
                    df = pd.DataFrame(data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # 格式化工作表
                    worksheet = writer.sheets[sheet_name]
                    
                    # 自动调整列宽
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # 设置标题行样式
                    if len(df) > 0:
                        for cell in worksheet[1]:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center')
                            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            return file_path
            
        except Exception as e:
            raise Exception(f"保存多工作表Excel文件失败: {str(e)}")
    
    async def validate_excel_data(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        验证Excel数据
        
        Args:
            data: 要验证的数据
        
        Returns:
            验证结果
        """
        try:
            if not data:
                return {
                    "valid": False,
                    "errors": ["数据为空"],
                    "warnings": [],
                    "summary": {"total_rows": 0, "total_columns": 0}
                }
            
            errors = []
            warnings = []
            
            # 检查数据结构一致性
            first_row_keys = set(data[0].keys())
            for i, row in enumerate(data[1:], 1):
                row_keys = set(row.keys())
                if row_keys != first_row_keys:
                    errors.append(f"第{i+1}行的列结构与第1行不一致")
            
            # 检查空值
            for i, row in enumerate(data):
                for key, value in row.items():
                    if value is None or (isinstance(value, str) and value.strip() == ""):
                        warnings.append(f"第{i+1}行的'{key}'列为空")
            
            # 检查数据类型一致性
            column_types = {}
            for row in data:
                for key, value in row.items():
                    if value is not None and value != "":
                        value_type = type(value).__name__
                        if key not in column_types:
                            column_types[key] = set()
                        column_types[key].add(value_type)
            
            for column, types in column_types.items():
                if len(types) > 1:
                    warnings.append(f"列'{column}'包含多种数据类型: {', '.join(types)}")
            
            return {
                "valid": len(errors) == 0,
                "errors": errors,
                "warnings": warnings,
                "summary": {
                    "total_rows": len(data),
                    "total_columns": len(first_row_keys),
                    "columns": list(first_row_keys),
                    "column_types": {k: list(v) for k, v in column_types.items()}
                }
            }
            
        except Exception as e:
            return {
                "valid": False,
                "errors": [f"验证过程中出错: {str(e)}"],
                "warnings": [],
                "summary": {"total_rows": 0, "total_columns": 0}
            }
    
    async def get_excel_templates(self) -> List[Dict[str, Any]]:
        """
        获取Excel模板列表
        
        Returns:
            模板信息列表
        """
        try:
            templates = []
            
            if self.template_dir.exists():
                for template_file in self.template_dir.glob("*.xlsx"):
                    template_info = {
                        "name": template_file.stem,
                        "filename": template_file.name,
                        "path": str(template_file),
                        "size": template_file.stat().st_size,
                        "created": datetime.fromtimestamp(template_file.stat().st_ctime).isoformat(),
                        "modified": datetime.fromtimestamp(template_file.stat().st_mtime).isoformat()
                    }
                    
                    # 尝试读取模板结构
                    try:
                        template_data = await self.read_excel_file(template_file)
                        template_info["sheets"] = list(template_data["sheets"].keys())
                        template_info["columns"] = {
                            sheet: info["columns"] 
                            for sheet, info in template_data["sheets"].items()
                        }
                    except:
                        template_info["sheets"] = []
                        template_info["columns"] = {}
                    
                    templates.append(template_info)
            
            return templates
            
        except Exception as e:
            raise Exception(f"获取Excel模板失败: {str(e)}")
    
    async def list_excel_files(self, directory: Optional[Path] = None) -> List[Dict[str, Any]]:
        """
        列出Excel文件
        
        Args:
            directory: 目录路径，默认为输出目录
        
        Returns:
            Excel文件信息列表
        """
        try:
            search_dir = directory if directory else self.output_dir
            
            if not search_dir.exists():
                return []
            
            excel_files = []
            
            for file_path in search_dir.glob("*.xlsx"):
                file_info = {
                    "name": file_path.stem,
                    "filename": file_path.name,
                    "path": str(file_path),
                    "size": file_path.stat().st_size,
                    "created": datetime.fromtimestamp(file_path.stat().st_ctime).isoformat(),
                    "modified": datetime.fromtimestamp(file_path.stat().st_mtime).isoformat()
                }
                
                # 尝试读取文件基本信息
                try:
                    excel_data = await self.read_excel_file(file_path)
                    file_info["sheets"] = excel_data["metadata"]["sheet_names"]
                    file_info["total_sheets"] = excel_data["metadata"]["total_sheets"]
                except:
                    file_info["sheets"] = []
                    file_info["total_sheets"] = 0
                
                excel_files.append(file_info)
            
            # 按修改时间排序（最新的在前）
            excel_files.sort(key=lambda x: x["modified"], reverse=True)
            
            return excel_files
            
        except Exception as e:
            raise Exception(f"列出Excel文件失败: {str(e)}")
    
    async def convert_to_csv(self, excel_path: Path, sheet_name: Optional[str] = None) -> Path:
        """
        将Excel文件转换为CSV
        
        Args:
            excel_path: Excel文件路径
            sheet_name: 工作表名称，如果为None则转换第一个工作表
        
        Returns:
            CSV文件路径
        """
        try:
            if not excel_path.exists():
                raise FileNotFoundError(f"Excel文件不存在: {excel_path}")
            
            # 读取Excel文件
            if sheet_name:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(excel_path)
            
            # 生成CSV文件路径
            csv_filename = excel_path.stem + ".csv"
            csv_path = excel_path.parent / csv_filename
            
            # 保存为CSV
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            
            return csv_path
            
        except Exception as e:
            raise Exception(f"转换为CSV失败: {str(e)}")
    
    async def merge_excel_files(
        self, 
        file_paths: List[Path], 
        output_filename: str,
        merge_type: str = "sheets"  # "sheets" or "rows"
    ) -> Path:
        """
        合并多个Excel文件
        
        Args:
            file_paths: Excel文件路径列表
            output_filename: 输出文件名
            merge_type: 合并类型 - "sheets"(作为不同工作表) 或 "rows"(合并行)
        
        Returns:
            合并后的文件路径
        """
        try:
            if not file_paths:
                raise ValueError("文件路径列表为空")
            
            # 确保输出文件名有正确的扩展名
            if not output_filename.endswith('.xlsx'):
                output_filename += '.xlsx'
            
            output_path = self.output_dir / output_filename
            
            if merge_type == "sheets":
                # 作为不同工作表合并
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    for i, file_path in enumerate(file_paths):
                        df = pd.read_excel(file_path)
                        sheet_name = f"Sheet_{i+1}_{file_path.stem}"
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            elif merge_type == "rows":
                # 合并行
                all_dfs = []
                for file_path in file_paths:
                    df = pd.read_excel(file_path)
                    all_dfs.append(df)
                
                merged_df = pd.concat(all_dfs, ignore_index=True)
                merged_df.to_excel(output_path, index=False)
            
            else:
                raise ValueError(f"不支持的合并类型: {merge_type}")
            
            return output_path
            
        except Exception as e:
            raise Exception(f"合并Excel文件失败: {str(e)}")