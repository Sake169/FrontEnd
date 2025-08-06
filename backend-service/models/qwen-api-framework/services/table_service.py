import json
from typing import Dict
from pathlib import Path
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from core.client import QwenClient
from services.prompt_templates import PromptTemplates

# 初始化logger
logger = logging.getLogger(__name__)

class TableExtractionService:
    RECORD_TYPES = [
        "证券投资报备交易信息",
        "证券投资报备持仓信息",
        "场外基金报备交易信息",
        "场外基金报备持仓信息",
        "未上市股权报备交易信息",
        "未上市股权报备持仓信息"
    ]

    # 各类型字段映射表（确保与Excel模板列名一致）
    FIELD_MAPPING = {
        "证券投资报备交易信息": {
            "证券账号": "证券账号",
            "交易市场": "交易市场",
            "证券代码": "证券代码",
            "证券名称": "证券名称",
            "买卖方向": "买卖方向",
            "交易类型": "交易类型",
            "交易数量（股）": "交易数量（股）",
            "交易价格": "交易价格",
            "产品名称": "产品名称",
            "成交日期": "成交日期",
            "备注": "备注"
        },
        "证券投资报备持仓信息": {
            "证券账号": "证券账号",
            "交易市场": "交易市场",
            "证券代码": "证券代码",
            "证券名称": "证券名称",
            "持仓数量（股）": "持仓数量（股）",
            "持仓日期": "持仓日期",
            "备注": "备注"
        },
        "场外基金报备交易信息": {
            "产品代码": "产品代码",
            "产品名称": "产品名称",
            "买卖方向": "买卖方向",
            "场外基金投向": "场外基金投向",
            "交易份额": "交易份额",
            "交易日期": "交易日期",
            "备注": "备注"
        },
        "场外基金报备持仓信息": {
            "产品代码": "产品代码",
            "产品名称": "产品名称",
            "持仓份额": "持仓份额",
            "持仓日期": "持仓日期",
            "备注": "备注"
        },
        "未上市股权报备交易信息": {
            "企业名称": "企业名称",
            "统一社会信用代码": "统一社会信用代码",
            "交易方向": "交易方向",
            "认缴金额（万元）": "认缴金额（万元）",
            "实缴金额（万元）": "实缴金额（万元）",
            "持股比例（%）": "持股比例（%）",
            "交易日期": "交易日期",
            "备注": "备注",
         },
         "未上市股权报备持仓信息": {
            "企业名称": "企业名称",
            "统一社会信用代码": "统一社会信用代码",
            "交易方向": "交易方向",
            "认缴金额（万元）": "认缴金额（万元）",
            "实缴金额（万元）": "实缴金额（万元）",
            "持股比例（%）": "持股比例（%）",
            "持仓日期": "持仓日期",
            "备注": "备注",
         }
    }
    def __init__(self):
        self.client = QwenClient()
    
    def extract_typed_records(self, combined_md: str) -> Dict:
        """提取带类型识别的结构化数据"""
        prompt = PromptTemplates.get_typed_extraction_prompt(combined_md)
        
        response = self.client.chat_completion(
            messages=[{"role": "user", "content": prompt.content}],
            temperature=0.1,
            response_format={"type": "json_object"}
        )
        
        content = response['choices'][0]['message']['content']
        return self._validate_response(content)
    
    def _validate_response(self, content: str) -> Dict:
        """验证并解析响应"""
        try:
            if '```json' in content:
                content = content.split('```json')[1].split('```')[0]
            
            data = json.loads(content)
            self._check_record_types(data)
            return data
            
        except json.JSONDecodeError as e:
            logger.error(f"JSON解析失败: {str(e)}\n原始内容: {content[:200]}...")
            raise
        except Exception as e:
            logger.error(f"验证响应时发生错误: {str(e)}")
            raise
    
    def _check_record_types(self, data: Dict):
        """验证记录类型是否合法"""
        for file_data in data.get("files", []):
            for record in file_data.get("records", []):
                if record.get("record_type") not in self.RECORD_TYPES:
                    error_msg = f"非法记录类型: {record.get('record_type')}"
                    logger.error(error_msg)
                    raise ValueError(error_msg)

    def write_to_excel_template(self, data: Dict, template_path: Path, output_path: Path) -> bool:
        """将数据写入已有Excel模板，合并所有sheet内容到一个sheet中"""
        logger.info(f"开始写入Excel模板: {template_path}")
        
        # 加载模板文件
        wb = load_workbook(template_path)
        
        # 创建一个新的合并工作表，或使用第一个工作表
        merged_sheet_name = "合并数据"
        if merged_sheet_name in wb.sheetnames:
            # 删除已存在的合并工作表
            wb.remove(wb[merged_sheet_name])
        
        # 创建新的合并工作表
        merged_sheet = wb.create_sheet(merged_sheet_name, 0)
        
        # 按记录类型分组数据
        records_by_type = {}
        for file_data in data.get("files", []):
            for record in file_data.get("records", []):
                record_type = record["record_type"]
                if record_type not in records_by_type:
                    records_by_type[record_type] = []
                records_by_type[record_type].append(record)
        
        current_row = 1
        
        # 按记录类型顺序处理数据
        for record_type in self.RECORD_TYPES:
            if record_type not in records_by_type:
                continue
                
            records = records_by_type[record_type]
            if not records:
                continue
            
            # 写入记录类型标题行
            merged_sheet.cell(row=current_row, column=1, value=record_type)
            # 设置标题行样式
            title_cell = merged_sheet.cell(row=current_row, column=1)
            title_cell.font = Font(bold=True, size=14)
            title_cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            current_row += 1
            
            # 获取字段映射关系
            field_map = self.FIELD_MAPPING.get(record_type, {})
            
            if field_map:
                # 写入字段标题行
                col = 1
                for field_name in field_map.keys():
                    header_cell = merged_sheet.cell(row=current_row, column=col, value=field_name)
                    header_cell.font = Font(bold=True)
                    header_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    col += 1
                current_row += 1
                
                # 写入数据行
                for record in records:
                    col = 1
                    for json_key in field_map.values():
                        value = record["data"].get(json_key, "")
                        merged_sheet.cell(row=current_row, column=col, value=value)
                        col += 1
                    current_row += 1
            
            # 在不同记录类型之间添加空行
            current_row += 1
        
        # 调整列宽
        for col in merged_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            merged_sheet.column_dimensions[column].width = adjusted_width
        
        # 保存新文件
        wb.save(output_path)
        logger.info(f"成功保存Excel文件: {output_path}")
        return True